# Copyright (c) 2026 Cerid AI. All rights reserved.
# SPDX-License-Identifier: Apache-2.0

"""XLSX parser — sheet-aware row-as-element with header replay.

Workstream E Phase 2b.6. Closes the audit's "XLSX: cell coordinates
and merged-cell context lost" gap. Each non-empty data row of every
sheet becomes one :class:`XLSXRow` element with:

* ``text`` formatted as ``"col1: val1 | col2: val2 | ..."`` (same
  row-replay shape as CSV from Phase 2b.1) so column semantics
  embed alongside values
* ``metadata`` with ``sheet_name`` (in addition to CSV's
  ``row_idx`` + ``column_headers`` + ``cells``) so retrieval can
  filter by sheet without parsing the workbook

Library choice: ``openpyxl`` — already in this codebase's
``requirements.txt`` (used by the legacy parser at
``app/parsers/excel.py``). Pure Python, no native deps. The only
additions to the stack are this module + the row chunker strategy.

Phase 2b.6 ships single-cell rows; merged-cell handling is queued
for the same follow-up that adds Pro-tier ``unstructured.partition_xlsx``
support — merged cells in the wild often span across header /
section boundaries and need the layout-aware library to disambiguate.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from core.ingest.parsers import ParsedElement

logger = logging.getLogger("ai-companion.ingest.parsers.xlsx")


def _cell_to_text(value: Any) -> str:
    """Render an openpyxl cell value as a string suitable for embedding.

    openpyxl returns native Python types (int/float/datetime/bool)
    plus formula strings starting with ``=`` (when ``data_only=False``).
    Datetimes render as ISO 8601; everything else uses the natural
    ``str()`` repr. None / empty cells render as empty strings.
    """
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        # datetime / date / time — ISO 8601 is searchable
        return value.isoformat()
    return str(value).strip()


def parse_xlsx(path: str | Path) -> list[ParsedElement]:
    """Parse an `.xlsx` file into one ``XLSXRow`` element per data row.

    Args:
        path: Filesystem path to the workbook.

    Returns:
        A list of :class:`ParsedElement` dicts. Each element has:

        * ``text`` — ``"col1: val1 | col2: val2 | ..."``
        * ``element_type`` — ``"XLSXRow"``
        * ``metadata`` — ``{sheet_name, row_idx, column_headers, cells}``

        Empty workbooks, header-only sheets, and wholly-empty rows
        within sheets are skipped. Returns ``[]`` when the file
        contains no usable data anywhere.

    Raises:
        FileNotFoundError: when ``path`` doesn't exist.
        ImportError: when openpyxl isn't installed.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"XLSX not found: {p}")

    # Lazy import — keeps module loadable when openpyxl missing.
    from openpyxl import load_workbook

    elements: list[ParsedElement] = []
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — fall back to legacy parser
        logger.warning("xlsx_open_failed file=%s error=%s", p.name, exc)
        return []

    try:
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            # Treat the first non-empty row as the header
            header_idx = 0
            while header_idx < len(rows) and not any(
                _cell_to_text(c) for c in rows[header_idx]
            ):
                header_idx += 1
            if header_idx >= len(rows):
                continue
            headers = [_cell_to_text(c) for c in rows[header_idx]]
            # openpyxl pads every row to the worksheet's max-column
            # width, so a 2-wide header on a sheet whose data rows
            # are 4-wide arrives as ['a', 'b', '', '']. Trim trailing
            # empty headers down to the last non-empty cell so the
            # column-count is the operator-intended width.
            while headers and not headers[-1]:
                headers.pop()
            if not headers:
                continue
            for row_idx, row in enumerate(
                rows[header_idx + 1 :], start=header_idx + 2,
            ):
                cells = [_cell_to_text(c) for c in row]
                # Pad/truncate to header width
                if len(cells) < len(headers):
                    cells.extend([""] * (len(headers) - len(cells)))
                cells = cells[: len(headers)]
                if not any(cells):
                    continue
                text = " | ".join(
                    f"{h}: {v}" for h, v in zip(headers, cells)
                )
                elements.append(
                    {
                        "text": text,
                        "element_type": "XLSXRow",
                        "metadata": {
                            "sheet_name": sheet_name,
                            "row_idx": row_idx,
                            "column_headers": headers,
                            "cells": cells,
                        },
                    },
                )
    finally:
        wb.close()

    logger.info(
        "xlsx_parsed file=%s rows=%d sheets=%d",
        p.name, len(elements), len(wb.sheetnames),
    )
    return elements
