# Copyright (c) 2026 Cerid AI. All rights reserved.
# SPDX-License-Identifier: Apache-2.0

"""Tests for the XLSX parser + chunker (Workstream E Phase 2b.6)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.ingest.chunkers import chunk_elements
from core.ingest.chunkers.xlsx_strategy import xlsx_row_strategy
from core.ingest.parsers.xlsx_parser import parse_xlsx


def _build_xlsx(path: Path, sheets: dict[str, list[list]]) -> Path:
    """Build a multi-sheet workbook from {sheet_name: [[row], ...]}."""
    wb = Workbook()
    # Drop default sheet
    if wb.worksheets:
        wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def test_parse_xlsx_emits_one_element_per_data_row(tmp_path):
    p = _build_xlsx(
        tmp_path / "sales.xlsx",
        {
            "Sales": [
                ["Quarter", "Revenue", "Region"],
                ["Q1", 1500000, "EMEA"],
                ["Q2", 1800000, "APAC"],
                ["Q3", 2100000, "AMER"],
            ],
        },
    )
    elements = parse_xlsx(p)
    assert len(elements) == 3
    for el in elements:
        assert el["element_type"] == "XLSXRow"
        assert el["metadata"]["sheet_name"] == "Sales"
        assert el["metadata"]["column_headers"] == ["Quarter", "Revenue", "Region"]


def test_parse_xlsx_text_replays_headers(tmp_path):
    p = _build_xlsx(
        tmp_path / "sales.xlsx",
        {"Sales": [["Q", "Revenue"], ["Q1", 1500000]]},
    )
    elements = parse_xlsx(p)
    # Numeric values render as their string repr
    assert elements[0]["text"] == "Q: Q1 | Revenue: 1500000"


def test_parse_xlsx_handles_multi_sheet_workbooks(tmp_path):
    p = _build_xlsx(
        tmp_path / "multi.xlsx",
        {
            "Q1": [["region", "revenue"], ["EMEA", 100], ["APAC", 200]],
            "Q2": [["region", "revenue"], ["EMEA", 150]],
        },
    )
    elements = parse_xlsx(p)
    sheet_counts: dict[str, int] = {}
    for el in elements:
        s = el["metadata"]["sheet_name"]
        sheet_counts[s] = sheet_counts.get(s, 0) + 1
    assert sheet_counts == {"Q1": 2, "Q2": 1}


def test_parse_xlsx_skips_wholly_empty_rows(tmp_path):
    """openpyxl emits None-filled rows for blank lines — must be dropped."""
    p = _build_xlsx(
        tmp_path / "sparse.xlsx",
        {
            "Sales": [
                ["id", "name"],
                [1, "alice"],
                [None, None],
                [2, "bob"],
            ],
        },
    )
    elements = parse_xlsx(p)
    assert len(elements) == 2
    names = [el["metadata"]["cells"][1] for el in elements]
    assert names == ["alice", "bob"]


def test_parse_xlsx_skips_leading_blank_rows(tmp_path):
    """Real-world workbooks often have a banner row + blank above headers."""
    p = _build_xlsx(
        tmp_path / "banner.xlsx",
        {
            "Sales": [
                [None, None, None],
                [None, None, None],
                ["Quarter", "Revenue"],
                ["Q1", 100],
            ],
        },
    )
    elements = parse_xlsx(p)
    assert len(elements) == 1
    assert elements[0]["metadata"]["column_headers"] == ["Quarter", "Revenue"]


def test_parse_xlsx_pads_short_rows(tmp_path):
    p = _build_xlsx(
        tmp_path / "short.xlsx",
        {"S": [["a", "b", "c"], ["1", "2"]]},
    )
    elements = parse_xlsx(p)
    assert elements[0]["metadata"]["cells"] == ["1", "2", ""]


def test_parse_xlsx_truncates_long_rows(tmp_path):
    p = _build_xlsx(
        tmp_path / "long.xlsx",
        {"S": [["a", "b"], ["1", "2", "3", "4"]]},
    )
    elements = parse_xlsx(p)
    assert elements[0]["metadata"]["cells"] == ["1", "2"]


def test_parse_xlsx_empty_workbook_returns_empty(tmp_path):
    p = _build_xlsx(tmp_path / "empty.xlsx", {"Sheet1": []})
    assert parse_xlsx(p) == []


def test_parse_xlsx_header_only_returns_empty(tmp_path):
    p = _build_xlsx(tmp_path / "hdr.xlsx", {"S": [["a", "b", "c"]]})
    assert parse_xlsx(p) == []


def test_parse_xlsx_row_idx_one_indexed_with_header_offset(tmp_path):
    """row_idx is the 1-indexed worksheet row number (header is row 1)."""
    p = _build_xlsx(
        tmp_path / "idx.xlsx",
        {"S": [["a"], ["1"], ["2"], ["3"]]},
    )
    elements = parse_xlsx(p)
    assert [el["metadata"]["row_idx"] for el in elements] == [2, 3, 4]


def test_parse_xlsx_renders_datetime_as_iso(tmp_path):
    """Date / datetime cells should embed in a searchable form."""
    from datetime import datetime
    p = tmp_path / "dt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["created_at", "name"])
    ws.append([datetime(2026, 4, 28, 10, 30, 0), "alice"])
    wb.save(str(p))
    elements = parse_xlsx(p)
    assert "2026-04-28T10:30:00" in elements[0]["text"]


def test_parse_xlsx_raises_on_missing_file(tmp_path):
    with pytest.raises(FileNotFoundError):
        parse_xlsx(tmp_path / "missing.xlsx")


def test_parse_xlsx_swallows_corrupt_file(tmp_path):
    """Corrupt XLSX bytes return [] so the dispatcher falls through."""
    p = tmp_path / "corrupt.xlsx"
    p.write_bytes(b"PK-broken garbage that openpyxl will reject")
    assert parse_xlsx(p) == []


# ---------------------------------------------------------------------------
# Chunker strategy + dispatch
# ---------------------------------------------------------------------------


def test_xlsx_row_strategy_passes_through():
    el = {
        "text": "Quarter: Q1 | Revenue: 1500000",
        "element_type": "XLSXRow",
        "metadata": {
            "sheet_name": "Sales",
            "row_idx": 2,
            "column_headers": ["Quarter", "Revenue"],
            "cells": ["Q1", "1500000"],
        },
    }
    chunks = xlsx_row_strategy(el)
    assert len(chunks) == 1
    md = chunks[0]["metadata"]
    assert md["element_type"] == "XLSXRow"
    assert md["sheet_name"] == "Sales"
    # column_headers must reach the chunk metadata — that's the audit fix
    assert md["column_headers"] == ["Quarter", "Revenue"]


def test_chunk_elements_dispatches_xlsx_strategy(tmp_path):
    p = _build_xlsx(
        tmp_path / "sales.xlsx",
        {"Sales": [["q", "rev"], ["Q1", 100], ["Q2", 200]]},
    )
    elements = parse_xlsx(p)
    chunks = chunk_elements(elements)
    assert len(chunks) == 2
    # Headers and sheet_name reach every chunk
    for c in chunks:
        assert c["metadata"]["element_type"] == "XLSXRow"
        assert c["metadata"]["sheet_name"] == "Sales"
        assert c["metadata"]["column_headers"] == ["q", "rev"]
